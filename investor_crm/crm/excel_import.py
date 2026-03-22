"""Parse Excel rows into Investor field values."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Any

from django.db.models.functions import Lower

from .models import Intermediary, Investor, Office, Responsibility


@dataclass
class ImportBatchContext:
    """Caches DB rows and responsibility FKs for large imports."""

    pref_resp: Responsibility | None = None
    int_resp: Responsibility | None = None
    office_by_norm: dict[str, Office] = field(default_factory=dict)
    intermediary_by_norm: dict[str, Intermediary] = field(default_factory=dict)
    # lower(name) -> Investor instance (from DB prefetch or newly created)
    investor_by_norm: dict[str, Investor] = field(default_factory=dict)

    def ensure_responsibilities(self) -> None:
        if self.pref_resp is None:
            self.pref_resp = Responsibility.objects.filter(name='Prefequity').first()
        if self.int_resp is None:
            self.int_resp = Responsibility.objects.filter(name='Intermediary').first()

    def get_office(self, name: str) -> Office | None:
        if not name:
            return None
        key = name[:255].strip().lower()
        if not key:
            return None
        if key not in self.office_by_norm:
            obj, _ = Office.objects.get_or_create(name=name[:255].strip())
            self.office_by_norm[key] = obj
        return self.office_by_norm[key]

    def get_intermediary(self, name: str) -> Intermediary | None:
        if not name:
            return None
        key = name[:255].strip().lower()
        if not key:
            return None
        if key not in self.intermediary_by_norm:
            obj, _ = Intermediary.objects.get_or_create(name=name[:255].strip())
            self.intermediary_by_norm[key] = obj
        return self.intermediary_by_norm[key]

    def prefetch_investors(self, names_lower: set[str]) -> None:
        """One query for all investors whose lower(name) is in names_lower."""
        if not names_lower:
            return
        lu = list(names_lower)
        # Chunk to avoid huge IN clauses
        chunk_size = 500
        for i in range(0, len(lu), chunk_size):
            chunk = lu[i : i + chunk_size]
            qs = (
                Investor.objects.annotate(inv_nl=Lower('name'))
                .filter(inv_nl__in=chunk)
                .select_related('office', 'responsibility', 'intermediary')
            )
            for inv in qs:
                self.investor_by_norm[inv.inv_nl] = inv

    def get_or_create_investor_slot(self, name: str) -> tuple[Investor, str]:
        key = name.strip().lower()
        if key in self.investor_by_norm:
            inv = self.investor_by_norm[key]
            return inv, 'updated' if inv.pk else 'created'
        inv = Investor(name=name.strip())
        self.investor_by_norm[key] = inv
        return inv, 'created'


def _norm_header(s: str) -> str:
    return re.sub(r'\s+', ' ', (s or '').strip().lower())


def _build_header_map():
    """Map normalized header variants to canonical keys."""
    return {
        'investor': 'name',
        'type': 'type',
        'contact': 'principal_contact',
        'principal contact': 'principal_contact',
        'principal contact(s)': 'principal_contact',
        'website': 'website',
        'status': 'statuses',
        'ticket size (£m)': 'ticket_size',
        'ticket size': 'ticket_size',
        'ticket size (m)': 'ticket_size',
        'meeting / zoom': 'meeting',
        'meeting': 'meeting',
        'vdr': 'vdr_access',
        'office': 'office',
        'prefequity resp.': 'prefequity_owner',
        'prefequity resp': 'prefequity_owner',
        'prefequity': 'prefequity_owner',
        'intermediary': 'intermediary',
        'comment': 'about',
        'comments': 'about',
    }


def _parse_bool(val: Any) -> bool:
    if val is None or val == '':
        return False
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    return s in ('yes', 'y', 'true', '1', 'x', '✓', '✔')


def _parse_decimal(val: Any):
    if val is None or val == '':
        return None
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    s = str(val).strip().replace(',', '').replace('£', '')
    if not s:
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _normalize_website(val: str) -> str:
    s = (val or '').strip()
    if not s:
        return ''
    if s.startswith(('http://', 'https://')):
        return s
    if '.' in s:
        return f'https://{s}'
    return s


# label (lower) -> status value
def _status_label_to_value() -> dict[str, str]:
    out = {}
    for value, label in Investor.STATUS_CHOICES:
        out[label.lower()] = value
        out[value.lower()] = value
    return out


# display label (lower) -> type value
def _type_label_to_value() -> dict[str, str]:
    out = {}
    for value, label in Investor.TYPE_CHOICES:
        out[label.lower()] = value
        out[value.replace('_', ' ').lower()] = value
        out[value.lower()] = value
    return out


def _parse_statuses(raw: Any) -> list[str]:
    if raw is None or raw == '':
        return []
    s = str(raw).strip()
    if not s:
        return []
    mapping = _status_label_to_value()
    # Longer labels first so "confirmed for sma" matches before "confirmed"
    ordered_labels = sorted(mapping.keys(), key=len, reverse=True)
    parts = re.split(r'[,;/\n|]+', s)
    result = []
    for part in parts:
        key = part.strip().lower()
        if not key:
            continue
        v = mapping.get(key)
        if v is None:
            for lbl in ordered_labels:
                if key == lbl or (len(key) > 3 and (key in lbl or lbl in key)):
                    v = mapping[lbl]
                    break
        if v is None:
            continue
        if v not in result:
            result.append(v)
    return result


def _parse_type(raw: Any) -> str:
    if raw is None or raw == '':
        return ''
    s = str(raw).strip().lower()
    mapping = _type_label_to_value()
    if s in mapping:
        return mapping[s]
    for lbl, val in mapping.items():
        if s == lbl or s in lbl:
            return val
    return ''


def _parse_prefequity_owner(raw: Any) -> str:
    if raw is None or raw == '':
        return ''
    s = str(raw).strip().upper()
    valid = {'TD', 'NP', 'JCP', 'TB', 'PW'}
    if s in valid:
        return s
    # e.g. "TD, NP" take first
    for token in re.split(r'[,;/\s]+', s):
        t = token.strip().upper()
        if t in valid:
            return t
    return ''


def _header_to_key(header: str) -> str | None:
    hmap = _build_header_map()
    nk = _norm_header(header)
    if nk in hmap:
        return hmap[nk]
    nk2 = nk.replace('(', '').replace(')', '').replace('£', '').strip()
    for hm, ck in hmap.items():
        hm2 = hm.replace('(', '').replace(')', '').replace('£', '').strip()
        if nk2 == hm2:
            return ck
    return None


def map_row(cells: dict[str, Any]) -> dict[str, Any]:
    """Map a dict of header -> cell value to investor kwargs (FKs as names / codes)."""
    canon = {}
    for k, v in cells.items():
        key = _header_to_key(k)
        if key:
            canon[key] = v

    name = str(canon.get('name') or '').strip()
    if not name:
        raise ValueError('Investor name is required')

    statuses = _parse_statuses(canon.get('statuses'))
    if not statuses:
        statuses = ['target_fund_iii']

    out = {
        'name': name,
        'principal_contact': str(canon.get('principal_contact') or '').strip(),
        'website': _normalize_website(str(canon.get('website') or '')),
        'statuses': statuses,
        'type': _parse_type(canon.get('type')),
        'ticket_size': _parse_decimal(canon.get('ticket_size')),
        'meeting': _parse_bool(canon.get('meeting')),
        'vdr_access': _parse_bool(canon.get('vdr_access')),
        'about': str(canon.get('about') or '').strip(),
        'office_name': str(canon.get('office') or '').strip(),
        'intermediary_name': str(canon.get('intermediary') or '').strip(),
        'prefequity_owner': _parse_prefequity_owner(canon.get('prefequity_owner')),
    }
    return out


def apply_investor_row(data: dict[str, Any], ctx: ImportBatchContext) -> tuple[Investor, str]:
    """
    Create or update an Investor from mapped row data using batch context caches.
    Returns (investor, action) where action is 'created' or 'updated'.
    """
    ctx.ensure_responsibilities()
    name = data['name']
    inv, action = ctx.get_or_create_investor_slot(name)

    inv.principal_contact = data.get('principal_contact', '')
    inv.website = (data.get('website') or '')[:200]
    inv.statuses = data.get('statuses') or ['target_fund_iii']
    inv.type = data.get('type', '')
    inv.ticket_size = data.get('ticket_size')
    inv.meeting = data.get('meeting', False)
    inv.vdr_access = data.get('vdr_access', False)
    inv.about = data.get('about', '')

    office_name = data.get('office_name', '')
    if office_name:
        inv.office = ctx.get_office(office_name)
    else:
        inv.office = None

    inter_name = data.get('intermediary_name', '')
    pref_owner = data.get('prefequity_owner', '')

    inv.intermediary = None
    inv.prefequity_owner = ''
    inv.responsibility = None

    if inter_name:
        inv.responsibility = ctx.int_resp
        inv.intermediary = ctx.get_intermediary(inter_name)
    elif pref_owner and ctx.pref_resp:
        inv.responsibility = ctx.pref_resp
        inv.prefequity_owner = pref_owner

    inv.full_clean()
    inv.save()
    return inv, action


def import_workbook(file_obj) -> dict[str, Any]:
    """
    Read an xlsx file object, return dict with keys:
    rows_ok, rows_failed, skipped_duplicates, created, updated, errors
    """
    from io import BytesIO

    from openpyxl import load_workbook

    if hasattr(file_obj, 'read') and not hasattr(file_obj, 'seek'):
        raw = file_obj.read()
        file_obj = BytesIO(raw)
    elif hasattr(file_obj, 'seek'):
        file_obj.seek(0)

    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {
            'rows_ok': 0,
            'rows_failed': 0,
            'skipped_duplicates': 0,
            'created': 0,
            'updated': 0,
            'errors': [{'row': 0, 'message': 'Empty sheet'}],
        }

    headers = [str(c).strip() if c is not None else '' for c in rows[0]]
    investor_col_idx = 0
    for j, h in enumerate(headers):
        if _header_to_key(h) == 'name':
            investor_col_idx = j
            break

    def _investor_cell_value(r: tuple) -> str:
        if investor_col_idx >= len(r):
            return ''
        v = r[investor_col_idx]
        if v is None:
            return ''
        return str(v).strip()

    # First row per investor name only; skip blank investor column; count file duplicates
    seen_name_lower: set[str] = set()
    skipped_duplicates = 0
    work_queue: list[tuple[int, tuple]] = []

    for i, row in enumerate(rows[1:], start=2):
        inv_val = _investor_cell_value(row)
        if not inv_val:
            continue
        nk = inv_val.strip().lower()
        if nk in seen_name_lower:
            skipped_duplicates += 1
            continue
        seen_name_lower.add(nk)
        work_queue.append((i, row))

    ctx = ImportBatchContext()
    ctx.ensure_responsibilities()
    ctx.prefetch_investors(seen_name_lower)

    errors = []
    created = updated = 0
    rows_ok = rows_failed = 0

    for i, row in work_queue:
        cells = {}
        for j, h in enumerate(headers):
            if j < len(row):
                cells[h] = row[j]
        try:
            data = map_row(cells)
            _, action = apply_investor_row(data, ctx)
            rows_ok += 1
            if action == 'created':
                created += 1
            else:
                updated += 1
        except Exception as e:
            rows_failed += 1
            errors.append({'row': i, 'message': str(e)})

    return {
        'rows_ok': rows_ok,
        'rows_failed': rows_failed,
        'skipped_duplicates': skipped_duplicates,
        'created': created,
        'updated': updated,
        'errors': errors[:50],
    }
